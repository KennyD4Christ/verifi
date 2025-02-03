from django.core.mail import send_mail
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak, Flowable
from django.conf import settings
from reportlab.pdfgen import canvas
from io import BytesIO
from reportlab.lib.pagesizes import letter, A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import inch
import io
import logging
from django.core.files.base import ContentFile
import csv
from .models import ReportEntry, ReportFile, Report, ReportAccessLog
import xlsxwriter
from django.db.models import Sum, Avg, Count, F, ExpressionWrapper, DecimalField, IntegerField, Case, When, Q, Subquery, OuterRef
import ast
from tenacity import (
    retry,
    stop_after_attempt,
    wait_exponential,
    retry_if_exception_type,
    before_sleep_log
)
from smtplib import (
    SMTPException,
    SMTPServerDisconnected,
    SMTPConnectError,
    SMTPResponseException
)
import math
import csv
from django.utils.html import strip_tags
import operator as op
import statistics
from transactions.models import Transaction
from products.models import Product
from core.models import Customer, Order, OrderItem
from django.utils import timezone
from django.db.models.functions import TruncMonth, Coalesce
from decimal import Decimal
from datetime import datetime, timedelta
from django.utils.timezone import make_aware
from django.core.exceptions import ValidationError
from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY, TA_LEFT, TA_RIGHT
from django.template import Template, Context
from django.core.mail import EmailMultiAlternatives
from django.utils.html import strip_tags
from typing import Optional, Dict, Any, List
from dataclasses import dataclass
from email.utils import formataddr
import bleach
from premailer import transform
from django.contrib.auth import get_user_model
from openpyxl.styles import NamedStyle, PatternFill, Font, Alignment, Border, Side, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension
import pandas as pd
import numpy as np
from reportlab.graphics.shapes import Drawing, Line, String, Rect
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.charts.linecharts import HorizontalLineChart
from reportlab.graphics.charts.legends import Legend
from reportlab.graphics.charts.piecharts import Pie
from calendar import month_name
from core.utils.currency import currency_formatter


logger = logging.getLogger(__name__)


@dataclass
class EmailRecipient:
    email: str
    full_name: Optional[str] = None
    role: Optional[str] = None
    permissions: List[str] = None

    def __post_init__(self):
        if self.permissions is None:
            self.permissions = []

    @property
    def display_name(self) -> str:
        if self.full_name:
            return self.full_name
        return self.email.split('@')[0].title()

    @property
    def is_administrator(self) -> bool:
        return 'Administrator' in self.role

    @property
    def is_auditor(self) -> bool:
        return 'Auditor' in self.role

class ReportContentGenerator:
    @staticmethod
    def _generate_summary_section(report, recipient, report_data=None):
        data = report_data or generate_comprehensive_report(report, None, None)
        summary_parts = []

        if recipient.is_administrator:
            summary_parts.extend([
                "<h2>Administrative Overview</h2>",
                f"<p>Total Records: {data['total_records']}</p>",
                f"<p>Last Updated: {data['updated_at']}</p>",
                "<p>System Status: All metrics within expected ranges</p>"
            ])
        elif recipient.is_auditor:
            summary_parts.extend([
                "<h2>Audit Summary</h2>",
                f"<p>Audit Trail: {report.get_audit_trail()}</p>",
                f"<p>Compliance Status: {data['compliance_status']}</p>",
                "<p>Key Findings:</p>",
                "<ul><li>All compliance requirements met</li></ul>"
            ])
        else:
            summary_parts.extend([
                "<h2>Report Summary</h2>",
                f"<p>Period: {data['start_date']} to {data['end_date']}</p>",
                "<p>Key Highlights:</p>",
                "<ul><li>Regular report summary available</li></ul>"
            ])

        return "\n".join(summary_parts)

    @staticmethod
    def _generate_manager_insights(report, report_data=None):
        # Use provided report_data or generate new data
        data = report_data or generate_comprehensive_report(report, None, None)
        return f"""
        <h2>Management Insights</h2>
        <div class="insights-section">
            <h3>Key Performance Indicators</h3>
            <ul>
                <li>Revenue Growth: {data['revenue_growth']}%</li>
                <li>Operational Efficiency: {data['operational_efficiency']}%</li>
                <li>Resource Utilization: {data['resource_utilization']}%</li>
            </ul>
            <h3>Action Items</h3>
            <ul>
                <li>Review performance metrics</li>
                <li>Assess resource allocation</li>
            </ul>
        </div>
        """

    @staticmethod
    def _generate_executive_summary(report, report_data=None):
        # Use provided report_data or generate new data
        data = report_data or generate_comprehensive_report(report, None, None)
        return f"""
        <h2>Executive Overview</h2>
        <div class="executive-summary">
            <p class="highlight">Overall Performance: {data['performance_rating']}</p>
            <div class="metrics-grid">
                <div class="metric">
                    <h4>Financial Impact</h4>
                    <p>{data,['financial_impact']}</p>
                </div>
                <div class="metric">
                    <h4>Strategic Alignment</h4>
                    <p>{data['strategic_alignment']}</p>
                </div>
            </div>
        </div>
        """

class EmailSendingError(Exception):
    """Custom exception for email sending failures"""
    def __init__(self, message: str, attempts: int = 0):
        self.attempts = attempts
        super().__init__(f"{message} (Attempts made: {attempts})")

def get_retry_config(exception: Exception) -> Dict[str, int]:
    """Determine retry configuration based on exception type"""
    if isinstance(exception, SMTPServerDisconnected):
        return {'max_attempts': 5, 'min_wait': 2, 'max_wait': 8}
    elif isinstance(exception, SMTPResponseException):
        if exception.smtp_code == 421:  # Service not available
            return {'max_attempts': 3, 'min_wait': 10, 'max_wait': 30}
    return {'max_attempts': 3, 'min_wait': 4, 'max_wait': 10}

def log_retry_attempt(retry_state):
    """Log information about retry attempts"""
    exc = retry_state.outcome.exception()
    retry_config = get_retry_config(exc)
    
    if retry_state.attempt_number >= retry_config['max_attempts'] - 1:
        logger.warning(
            f"Final retry attempt. Error: {str(exc)}. "
            f"Total attempts: {retry_state.attempt_number + 1}"
        )
    else:
        logger.info(
            f"Retry attempt {retry_state.attempt_number + 1} of {retry_config['max_attempts']}. "
            f"Error: {str(exc)}"
        )

@retry(
    stop=stop_after_attempt(3),
    wait=wait_exponential(multiplier=1, min=4, max=10),
    before_sleep=before_sleep_log(logger, logging.WARNING),
    retry=retry_if_exception_type((SMTPException, ConnectionError))
)
def send_enhanced_report_email(
    report,
    recipient: EmailRecipient,
    include_summary: bool = True,
    include_charts: bool = True,
    start_date_str: Optional[str] = None,
    end_date_str: Optional[str] = None,
    pdf_content: Optional[bytes] = None
) -> None:
    """
    Send enhanced report email with retry mechanism and improved error handling.
    """
    try:
        if not start_date_str:
            start_date_str = report.start_date.strftime('%Y-%m-%d') if report.start_date else report.created_at.strftime('%Y-%m-%d')
        if not end_date_str:
            end_date_str = report.end_date.strftime('%Y-%m-%d') if report.end_date else timezone.now().strftime('%Y-%m-%d')

        logger.info(f"Processing report email for period: {start_date_str} to {end_date_str}")

        report_data = generate_comprehensive_report(report, start_date_str, end_date_str)
        logger.info(f"Email report data for {report.id}: Total Revenue: ${report_data['financial_overview']['total_revenue']}")

        email_components = prepare_email_components(
            report=report,
            recipient=recipient,
            include_summary=include_summary,
            include_charts=include_charts,
            report_data=report_data
        )
        
        if pdf_content is None:
            with generate_pdf_report(report, start_date_str=start_date_str, end_date_str=end_date_str) as pdf_file:
                pdf_content = pdf_file.read()

        attempts = send_email_with_components(email_components, recipient, pdf_content)
        return attempts
        
        log_email_success(
            report=report,
            recipient=recipient,
            report_data=report_data,
            include_summary=include_summary,
            include_charts=include_charts,
            attempts=attempts
        )
        
        logger.info(f"Report email sent successfully to {recipient.email}")

    except Exception as e:
        logger.error(f"Failed to send report email to {recipient.email}: {str(e)}")
        raise EmailSendingError(str(e))

def prepare_email_components(
    report,
    recipient: EmailRecipient,
    include_summary: bool,
    include_charts: bool,
    report_data: Dict[str, Any]
) -> Dict[str, Any]:
    """
    Prepare all components needed for the email.
    """
    try:
        # Generate PDF report
        pdf_content = generate_pdf_with_retry(report, report_data)

        # Get email template
        email_template = Report.get_email_template()
        if not email_template:
            raise ValueError("Email template configuration is missing")

        # Generate email content
        content_generator = ReportContentGenerator()
        email_content = _generate_email_content(
            report=report,
            recipient=recipient,
            include_summary=include_summary,
            include_charts=include_charts,
            content_generator=content_generator,
            report_data=report_data
        )

        # Process financial data for context
        financial_overview = report_data['financial_overview'].copy()
        for key in ['total_revenue', 'net_profit', 'operating_expenses', 'cost_of_services']:
            if key in financial_overview:
                value = parse_number(financial_overview[key])
                financial_overview[key] = format_number(value, include_suffix=True)

        # Process performance metrics
        performance_metrics = report_data['performance_metrics'].copy()
        if 'order_performance' in performance_metrics:
            order_perf = performance_metrics['order_performance']
            if 'average_order_value' in order_perf:
                value = parse_number(order_perf['average_order_value'])
                order_perf['average_order_value'] = format_number(value, include_suffix=True)

        # Prepare context
        context = {
            **report.generate_email_context(recipient_name=recipient.display_name),
            'email_content': email_content,
            'period_start': report_data['report_metadata']['period_start'],
            'period_end': report_data['report_metadata']['period_end'],
            'financial_overview': financial_overview,
            'performance_metrics': performance_metrics,
            'inventory_insights': report_data['inventory_insights']
        }

        return {
            'pdf_content': pdf_content,
            'email_template': email_template,
            'context': context,
            'report_name': report.name,
            'report_data': report_data
        }
    except Exception as e:
        logger.error(f"Error preparing email components: {str(e)}")
        raise

@retry(
    stop=stop_after_attempt(3),
    wait=wait_exponential(multiplier=1, min=4, max=10),
    before_sleep=log_retry_attempt
)
def generate_pdf_with_retry(report, report_data: dict) -> bytes:
    """Generate PDF report with retry mechanism"""
    try:
        with generate_pdf_report(report, report_data=report_data) as pdf_file:
            return pdf_file.read()
    except Exception as e:
        logger.error(f"PDF generation failed: {str(e)}")
        raise

@retry(
    stop=stop_after_attempt(3),
    wait=wait_exponential(multiplier=1, min=4, max=10),
    before_sleep=log_retry_attempt
)
def send_email_with_components(
    email_components: Dict[str, Any],
    recipient: EmailRecipient,
    pdf_content: bytes
) -> int:
    """Send email with retry mechanism and return number of attempts"""
    attempts = 0
    
    try:
        html_template = Template(email_components['email_template'].template_content)
        html_content = transform(html_template.render(Context(email_components['context'])))
        text_content = strip_tags(html_content)
        
        period_str = (
            f"{email_components['report_data']['report_metadata']['period_start']} to "
            f"{email_components['report_data']['report_metadata']['period_end']}"
        )
        
        msg = EmailMultiAlternatives(
            subject=f"Report: {email_components['report_name']} - {period_str}",
            body=text_content,
            from_email=formataddr((settings.COMPANY_NAME, settings.DEFAULT_FROM_EMAIL)),
            to=[formataddr((recipient.display_name, recipient.email))]
        )
        
        msg.attach_alternative(html_content, "text/html")
        msg.attach(
            f"{email_components['report_name']}_report.pdf",
            pdf_content,
            'application/pdf'
        )
        
        msg.send(fail_silently=False)
        return attempts
        
    except SMTPResponseException as e:
        attempts += 1
        logger.warning(
            f"SMTP response error (code: {e.smtp_code}) on attempt {attempts}: {str(e)}"
        )
        raise
    except SMTPException as e:
        attempts += 1
        logger.warning(f"SMTP error on attempt {attempts}: {str(e)}")
        raise
    except Exception as e:
        attempts += 1
        logger.error(f"Unexpected error on attempt {attempts}: {str(e)}")
        raise

def log_email_success(
    report,
    recipient: EmailRecipient,
    report_data: dict,
    include_summary: bool,
    include_charts: bool,
    attempts: int
) -> None:
    """Log successful email sending with metadata"""
    # Get user if exists
    User = get_user_model()
    try:
        user = User.objects.get(email=recipient.email)
    except User.DoesNotExist:
        user = None
    
    # Create access log
    ReportAccessLog.objects.create(
        report=report,
        user=user,
        action='email_sent',
        metadata={
            'recipient_email': recipient.email,
            'recipient_role': recipient.role,
            'included_summary': include_summary,
            'included_charts': include_charts,
            'pdf_attached': True,
            'period_start': report_data['report_metadata']['period_start'],
            'period_end': report_data['report_metadata']['period_end'],
            'data_timestamp': timezone.now().isoformat(),
            'sending_attempts': attempts
        }
    )
    
    if attempts > 1:
        logger.warning(f"Email sent successfully after {attempts} attempts to {recipient.email}")

def _generate_email_content(
    report,
    recipient: EmailRecipient,
    include_summary: bool,
    include_charts: bool,
    content_generator: ReportContentGenerator,
    report_data: dict
) -> str:
    content_parts = []

    financial_data = report_data['financial_overview']
    performance_data = report_data['performance_metrics']

    # Parse financial values
    total_revenue = parse_number(financial_data['total_revenue'])
    net_profit = parse_number(financial_data['net_profit'])
    operating_expenses = parse_number(financial_data['operating_expenses'])
    cost_of_services = parse_number(financial_data['cost_of_services'])
    avg_order_value = parse_number(performance_data['order_performance']['average_order_value'])

    if recipient.is_administrator:
        content_parts.extend([
            "<h2>Administrative Overview</h2>",
            f"<p>Period: {report_data['report_metadata']['period_start']} to {report_data['report_metadata']['period_end']}</p>",
            f"<p>Total Revenue: {format_number(total_revenue, include_suffix=True)}</p>",
            f"<p>Net Profit: {format_number(net_profit, include_suffix=True)}</p>",
            f"<p>Total Orders: {performance_data['order_performance']['total_orders']}</p>"
        ])
    elif recipient.is_auditor:
        content_parts.extend([
            "<h2>Audit Summary</h2>",
            f"<p>Financial Period: {report_data['report_metadata']['period_start']} to {report_data['report_metadata']['period_end']}</p>",
            f"<p>Revenue Verified: {format_number(total_revenue, include_suffix=True)}</p>",
            f"<p>Expenses Verified: {format_number(operating_expenses, include_suffix=True)}</p>",
            f"<p>Cost of Services: {format_number(cost_of_services, include_suffix=True)}</p>"
        ])
    else:
        content_parts.extend([
            "<h2>Report Summary</h2>",
            f"<p>Period: {report_data['report_metadata']['period_start']} to {report_data['report_metadata']['period_end']}</p>",
            "<p>Performance Overview:</p>",
            f"<p>Average Order Value: {format_number(avg_order_value, include_suffix=True)}</p>"
        ])

    return "\n".join(content_parts)

def validate_date_range(start_date_str: Optional[str], end_date_str: Optional[str]) -> tuple:
    """
    Validates and processes date range inputs, handling various scenarios and ensuring proper timezone awareness.
    Returns a tuple of validated start and end datetime objects.
    """
    try:
        current_time = timezone.now()

        # Handle cases where dates aren't provided
        if not start_date_str and not end_date_str:
            # Default to last 30 days if no dates provided
            end_date = current_time
            start_date = end_date - timedelta(days=30)
            return start_date, end_date

        # Process provided dates
        if start_date_str:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
            start_date = make_aware(start_date)
        else:
            # If only end date provided, default start date to 30 days before
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
            end_date = make_aware(end_date)
            start_date = end_date - timedelta(days=30)

        if end_date_str:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
            end_date = make_aware(
                datetime.combine(end_date.date(), datetime.max.time())
            )
        else:
            # If only start date provided, default end date to start date + 30 days
            end_date = min(start_date + timedelta(days=30), current_time)

        # Validate date range
        if end_date < start_date:
            raise ValidationError("End date must be after start date.")

        # Ensure end date doesn't exceed current time
        if end_date > current_time:
            end_date = current_time

        return start_date, end_date

    except ValueError:
        raise ValidationError("Invalid date format. Please use YYYY-MM-DD format.")
    except Exception as e:
        raise ValidationError(f"Error validating date range: {str(e)}")

def format_number(value, include_suffix=False):
    """
    Format numbers with thousand separators and optional K/M suffix.
    
    Args:
        value (float/Decimal): The number to format
        include_suffix (bool): Whether to include K/M suffix for large numbers
        
    Returns:
        str: Formatted number string
    """
    try:
        value = float(value)
        if include_suffix:
            if value >= 1_000_000:
                return f"{value/1_000_000:,.2f}M"
            elif value >= 1_000:
                return f"{value/1_000:,.2f}K"
        return f"{value:,.2f}"
    except (TypeError, ValueError):
        return "0.00"

def parse_number(value_str):
    """
    Parse a formatted number string back to float.
    
    Args:
        value_str (str): The formatted number string or number
        
    Returns:
        float: Parsed number value
    """
    if isinstance(value_str, (int, float, Decimal)):
        return float(value_str)
    try:
        # Handle strings with K/M suffixes
        value_str = str(value_str).strip().upper()
        multiplier = 1
        if value_str.endswith('M'):
            multiplier = 1_000_000
            value_str = value_str[:-1]
        elif value_str.endswith('K'):
            multiplier = 1_000
            value_str = value_str[:-1]
        return float(value_str.replace(',', '')) * multiplier
    except (ValueError, AttributeError):
        return 0.0

def generate_comprehensive_report(report, start_date_str: Optional[str], end_date_str: Optional[str]) -> dict:
    try:
        if start_date_str and end_date_str:
            # Use validated date range when dates are provided
            start_date, end_date = validate_date_range(start_date_str, end_date_str)
        else:
            # Handle cases where one or both dates are missing
            if start_date_str:
                # If only start date provided, validate it and set end date to current time
                temp_start, end_date = validate_date_range(start_date_str, timezone.now().strftime('%Y-%m-%d'))
                start_date = temp_start
            elif end_date_str:
                # If only end date provided, validate it and set start date to 30 days prior
                end_date_obj = datetime.strptime(end_date_str, '%Y-%m-%d')
                start_date_obj = end_date_obj - timedelta(days=30)
                start_date, end_date = validate_date_range(
                    start_date_obj.strftime('%Y-%m-%d'),
                    end_date_str
                )
            else:
                # If no dates provided, use custom default range based on report parameters
                if hasattr(report, 'reporting_period') and report.reporting_period:
                    # Use report's configured period if available
                    start_date = report.created_at
                    end_date = start_date + timedelta(days=report.reporting_period)
                    # Ensure end date doesn't exceed current time
                    if end_date > timezone.now():
                        end_date = timezone.now()
                else:
                    # Default to report creation date and current time
                    start_date = report.created_at
                    end_date = timezone.now()

        logger.info(f"Generating report for period: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}")

        def _analyze_income_breakdown(income_transactions):
            categorized = income_transactions.exclude(
                Q(category__isnull=True) | Q(category='')
            ).values(
                'category'
            ).annotate(
                total_amount=Sum('amount'),
                transaction_count=Count('id')
            ).order_by('-total_amount')

            category_mapping = dict(Transaction.CATEGORY_CHOICES)

            breakdown = []
            for item in categorized:
                category_name = category_mapping.get(item['category'])
                if category_name:
                    breakdown.append({
                        'category': category_name,
                        'total_amount': format_number(item['total_amount']),
                        'transaction_count': item['transaction_count']
                    })

            uncategorized = income_transactions.filter(
                Q(category__isnull=True) | Q(category='')
            ).aggregate(
                total_amount=Sum('amount'),
                transaction_count=Count('id')
            )

            if uncategorized['total_amount']:
                breakdown.append({
                    'category': 'Uncategorized Income',
                    'total_amount': format_number(uncategorized['total_amount']),
                    'transaction_count': uncategorized['transaction_count']
                })

            return breakdown

        def _analyze_expense_breakdown(expense_transactions):
            expense_data = expense_transactions.values('category').annotate(
                total_amount=Sum('amount'),
                transaction_count=Count('id')
            ).order_by('-total_amount')

            return [{
                'category': item['category'],
                'total_amount': format_number(item['total_amount']),
                'transaction_count': item['transaction_count']
            } for item in expense_data]

        def _calculate_monthly_cash_flow(transactions):
            cash_flow = transactions.annotate(
                month=TruncMonth('date')
            ).values('month', 'transaction_type').annotate(
                total_amount=Sum('amount')
            ).order_by('month')

            return [{
                'month': item['month'],
                'transaction_type': item['transaction_type'],
                'total_amount': format_number(item['total_amount'])
            } for item in cash_flow]

        def generate_inventory_insights():
            date_filter = Q(
                order_items__order__order_date__range=[start_date, end_date],
                order_items__order__status__in=['delivered', 'shipped']
            )
            inventory_data = Product.objects.filter(is_active=True).annotate(
                total_stock=Sum('stock'),
                total_value=ExpressionWrapper(
                    F('stock') * F('price'),
                    output_field=DecimalField()
                ),
                total_sales=Coalesce(
                    Sum(
                        Case(
                            When(
                                order_items__order__status__in=['delivered', 'shipped'],
                                then='order_items__quantity'
                            ),
                            default=0,
                            output_field=IntegerField(),
                        ),
                        filter=date_filter
                    ),
                    0,
                    output_field=IntegerField()
                ),
                total_revenue=Coalesce(
                    Sum(
                        Case(
                            When(
                                order_items__order__status__in=['delivered', 'shipped'],
                                then=F('order_items__quantity') * F('order_items__unit_price')
                            ),
                            default=0,
                            output_field=DecimalField(max_digits=10, decimal_places=2)
                        ),
                        filter=date_filter
                    ),
                    0,
                    output_field=DecimalField(max_digits=10, decimal_places=2)
                )
            )

            low_stock_products = inventory_data.filter(
                stock__lt=F('low_stock_threshold')
            ).values(
                'name',
                'stock',
                'low_stock_threshold',
                'total_sales'
            ).order_by('stock')

            total_stock_value = inventory_data.aggregate(
                total=Sum(F('stock') * F('price'))
            )['total'] or Decimal('0.00')

            return {
                'total_product_count': inventory_data.count(),
                'total_stock_value': format_number(total_stock_value),
                'inventory_status': {
                    'status': 'Healthy' if not low_stock_products else 'Attention Required',
                    'low_stock_count': low_stock_products.count()
                },
                'low_stock_products': list(low_stock_products),
                'top_selling_products': [{
                    'name': product['name'],
                    'total_sales': product['total_sales'],
                    'total_revenue': format_number(product['total_revenue'])
                } for product in inventory_data.filter(
                    Q(total_sales__gt=0) | Q(total_revenue__gt=0)
                ).order_by('-total_sales')[:10].values(
                    'name',
                    'total_sales',
                    'total_revenue'
                )]
            }

        def generate_performance_metrics():
            date_filter = Q(orders__order_date__range=[start_date, end_date])
            customer_metrics = Customer.objects.annotate(
                total_orders=Count('orders'),
                total_spend=Coalesce(
                    Sum(
                        F('orders__items__quantity') * F('orders__items__unit_price'),
                        filter=date_filter & Q(orders__status__in=['delivered', 'shipped'])
                    ),
                    Decimal('0.00'),
                    output_field=DecimalField(max_digits=10, decimal_places=2)
                )
            )

            order_performance = Order.objects.filter(
                order_date__range=[start_date, end_date]
            ).aggregate(
                total_orders=Count('id'),
                average_order_value=Coalesce(
                    Avg(
                        Subquery(
                            Order.objects.filter(
                                id=OuterRef('id')
                            ).annotate(
                                order_total=Sum(F('items__quantity') * F('items__unit_price'))
                            ).values('order_total')[:1]
                        )
                    ),
                    Decimal('0.00'),
                    output_field=DecimalField(max_digits=10, decimal_places=2)
                ),
                total_revenue=Coalesce(
                    Sum(F('items__quantity') * F('items__unit_price')),
                    Decimal('0.00'),
                    output_field=DecimalField(max_digits=10, decimal_places=2)
                )
            )

            def _calculate_sales_trend():
                sales_data = Order.objects.filter(
                    order_date__range=[start_date, end_date]
                ).annotate(
                    month=TruncMonth('order_date')
                ).values('month').annotate(
                    total_sales=Coalesce(
                        Sum(F('items__quantity') * F('items__unit_price')),
                        Decimal('0.00'),
                        output_field=DecimalField(max_digits=10, decimal_places=2)
                    )
                ).order_by('month')

                return [{
                    'month': item['month'],
                    'total_sales': format_number(item['total_sales'])
                } for item in sales_data]

            return {
                'customer_metrics': {
                    'total_customers': customer_metrics.count(),
                    'average_orders_per_customer': float(
                        statistics.mean([c.total_orders for c in customer_metrics] or [0])
                    ),
                    'top_customers': [{
                        'id': customer['id'],
                        'first_name': customer['first_name'],
                        'last_name': customer['last_name'],
                        'email': customer['email'],
                        'total_orders': customer['total_orders'],
                        'total_spend': format_number(customer['total_spend'])
                    } for customer in customer_metrics.order_by('-total_spend')[:5].values(
                        'id', 'first_name', 'last_name', 'email', 'total_orders', 'total_spend'
                    )]
                },
                'order_performance': {
                    'total_orders': order_performance['total_orders'],
                    'average_order_value': format_number(order_performance['average_order_value']),
                    'total_revenue': format_number(order_performance['total_revenue'])
                },
                'sales_trend': _calculate_sales_trend()
            }

        transactions = Transaction.objects.filter(
            date__range=[start_date, end_date]
        )

        revenue = transactions.filter(
            transaction_type='income'
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')

        cost_of_services = transactions.filter(
            transaction_type='cost_of_services'
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')

        operating_expenses = transactions.filter(
            transaction_type='expense'
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')

        net_profit = revenue - cost_of_services - operating_expenses

        return {
            'report_metadata': {
                'period_start': start_date.strftime('%Y-%m-%d'),
                'period_end': end_date.strftime('%Y-%m-%d')
            },
            'financial_overview': {
                'total_revenue': format_number(revenue),
                'cost_of_services': format_number(cost_of_services),
                'operating_expenses': format_number(operating_expenses),
                'net_profit': format_number(net_profit),
                'income_breakdown': _analyze_income_breakdown(
                    transactions.filter(transaction_type='income')
                ),
                'expense_breakdown': _analyze_expense_breakdown(
                    transactions.filter(transaction_type__in=['expense', 'cost_of_services'])
                ),
                'monthly_cash_flow': _calculate_monthly_cash_flow(transactions)
            },
            'inventory_insights': generate_inventory_insights(),
            'performance_metrics': generate_performance_metrics()
        }

    except ValueError as e:
        raise ValueError(f"Error generating report: {str(e)}")

def create_revenue_trend_chart(financial_data, width, height):
    drawing = Drawing(width, height)
    chart = HorizontalLineChart()

    chart.x = 60
    chart.y = 50
    chart.width = width - 120
    chart.height = height - 100

    try:
        revenue_data = financial_data.get('monthly_revenue', [])

        if not revenue_data:
            chart.data = [[0, 0]]
            chart.categoryAxis.categoryNames = ['No Data', '']
            chart.valueAxis.valueMin = 0
            chart.valueAxis.valueMax = 100
            drawing.add(chart)
            return drawing

        def sort_key(x):
            date = datetime.strptime(f"{x['month']} {x['year']}", '%B %Y')
            return date

        sorted_data = sorted(revenue_data, key=sort_key)

        if len(sorted_data) == 1:
            sorted_data.append(sorted_data[0].copy())

        revenue_values = [float(entry['revenue']) for entry in sorted_data]

        months = []
        prev_year = None
        for entry in sorted_data:
            current_year = entry['year']
            if prev_year is None or current_year != prev_year:
                months.append(f"{entry['month'][:3]} {str(current_year)[2:]}")
            else:
                months.append(entry['month'][:3])
            prev_year = current_year

        if len(revenue_values) >= 3:
            window_size = 3
            moving_averages = []
            for i in range(len(revenue_values)):
                if i < window_size - 1:
                    moving_averages.append(revenue_values[i])
                else:
                    avg = sum(revenue_values[i-(window_size-1):i+1]) / window_size
                    moving_averages.append(avg)
            chart.data = [revenue_values, moving_averages]
        else:
            chart.data = [revenue_values]

        chart.categoryAxis.categoryNames = months

        min_value = min(revenue_values) if revenue_values else 0
        max_value = max(revenue_values) if revenue_values else 100
        padding = (max_value - min_value) * 0.15 if max_value > min_value else 10

        chart.valueAxis.valueMin = float(min_value - padding)
        chart.valueAxis.valueMax = float(max_value + padding)

        chart.categoryAxis.labels.boxAnchor = 'ne'
        chart.categoryAxis.labels.angle = 30
        chart.categoryAxis.labels.fontSize = 8
        chart.categoryAxis.labels.fontName = 'Helvetica'
        chart.categoryAxis.labels.dy = -8
        chart.categoryAxis.strokeWidth = 1
        chart.categoryAxis.strokeColor = colors.HexColor('#424242')

        x_axis_label = String(
            width / 2, 20, 'Month',
            fontSize=10, fontName='Helvetica',
            textAnchor='middle'
        )
        drawing.add(x_axis_label)

        chart.valueAxis.labels.fontSize = 8
        chart.valueAxis.labels.fontName = 'Helvetica'
        chart.valueAxis.strokeWidth = 1
        chart.valueAxis.strokeColor = colors.HexColor('#424242')

        y_axis_label = String(
            15, height / 2, 'Revenue',
            fontSize=10, fontName='Helvetica',
            textAnchor='middle',
            angle=90
        )
        drawing.add(y_axis_label)

        chart.valueAxis.labelTextFormat = lambda x: format_number(x, include_suffix=True)

        chart.lines[0].strokeWidth = 3
        chart.lines[0].strokeColor = colors.HexColor('#1a237e')
        if len(chart.data) > 1:
            chart.lines[1].strokeWidth = 2
            chart.lines[1].strokeColor = colors.HexColor('#43a047')

        for i, value in enumerate(revenue_values):
            label = String(
                chart.x + (i + 0.5) * (chart.width / len(revenue_values)),
                chart.y + chart.height * (value - chart.valueAxis.valueMin) /
                (chart.valueAxis.valueMax - chart.valueAxis.valueMin),
                format_number(value, include_suffix=True),
                fontSize=8,
                fontName='Helvetica',
                textAnchor='middle'
            )
            drawing.add(label)

        legend_y = height - 25

        drawing.add(Line(
            width - 220, legend_y, width - 190, legend_y,
            strokeWidth=3,
            strokeColor=colors.HexColor('#1a237e')
        ))
        drawing.add(String(
            width - 180, legend_y,
            'Actual Revenue',
            fontSize=8,
            fontName='Helvetica'
        ))

        if len(chart.data) > 1:
            drawing.add(Line(
                width - 120, legend_y, width - 90, legend_y,
                strokeWidth=2,
                strokeColor=colors.HexColor('#43a047')
            ))
            drawing.add(String(
                width - 80, legend_y,
                '3-Month Trend',
                fontSize=8,
                fontName='Helvetica'
            ))

    except Exception as e:
        print(f"Error creating revenue chart: {str(e)}")
        chart.data = [[0, 0]]
        chart.categoryAxis.categoryNames = ['Error', '']
        chart.valueAxis.valueMin = 0
        chart.valueAxis.valueMax = 100

    drawing.add(chart)
    return drawing

def create_category_pie_chart(data, width, height):
    drawing = Drawing(width, height)
    pie = Pie()

    pie.width = min(width, height) - 120
    pie.height = min(width, height) - 120
    pie.x = (width - pie.width) / 2
    pie.y = (height - pie.height) / 2

    valid_categories = []
    for category in data:
        if (category.get('total_amount', 0) > 0 and
                category.get('category') and
                isinstance(category['category'], str)):
            valid_categories.append(category)

    if not valid_categories:
        pie.data = [100]
        pie.labels = ['No Data']
    else:
        pie.data = [float(str(category['total_amount']).replace(',', ''))
                   for category in valid_categories]
        pie.labels = [category['category'] for category in valid_categories]

    pie.slices.strokeWidth = 1
    pie.slices.strokeColor = colors.white

    chart_colors = [
        colors.HexColor('#1976d2'),
        colors.HexColor('#26c6da'),
        colors.HexColor('#66bb6a'),
        colors.HexColor('#ffa726'),
        colors.HexColor('#ff7043'),
        colors.HexColor('#ab47bc'),
        colors.HexColor('#ec407a'),
    ]
    for i in range(len(pie.data)):
        pie.slices[i].fillColor = chart_colors[i % len(chart_colors)]

    pie.sideLabels = True
    pie.simpleLabels = False
    pie.slices.popout = 10

    total = sum(pie.data)
    if total > 0:
        percentages = [(value / total) * 100 for value in pie.data]
        pie.labels = [
            f'{label} ({format_number(value)}) - {perc:.1f}%'
            for label, value, perc in zip(pie.labels, pie.data, percentages)
        ]

    drawing.add(pie)
    return drawing

def create_inventory_bar_chart(data, width, height):
    drawing = Drawing(width, height)

    chart = VerticalBarChart()
    chart.x = 50
    chart.y = 50
    chart.width = width - 100
    chart.height = height - 150

    chart.data = [data['stock_levels']]
    chart.categoryAxis.categoryNames = data['product_categories']
    chart.categoryAxis.labels.boxAnchor = 'ne'
    chart.categoryAxis.labels.angle = 30

    chart.bars[0].fillColor = colors.HexColor('#1e88e5')
    chart.bars[0].strokeWidth = 0.5
    chart.bars[0].strokeColor = colors.HexColor('#1565c0')

    chart.valueAxis.gridStrokeWidth = 0.5
    chart.valueAxis.gridStrokeColor = colors.HexColor('#e0e0e0')

    chart.valueAxis.valueMin = 0
    chart.valueAxis.labels.fontSize = 10
    chart.valueAxis.labelTextFormat = lambda x: format_number(x, include_suffix=True)
    chart.categoryAxis.labels.boxAnchor = 'ne'
    chart.categoryAxis.labels.angle = 45
    chart.categoryAxis.labels.fontSize = 9
    chart.categoryAxis.labels.fillColor = colors.HexColor('#424242')

    legend_label = String(
        width - 150, height - 50, "Stock Value",
        fontSize=12,
        fillColor=colors.HexColor('#1e88e5')
    )
    drawing.add(legend_label)

    drawing.add(chart)
    return drawing

def calculate_growth_rate(current_value, previous_value):
    if current_value is None or previous_value is None:
        return 0.0

    try:
        if isinstance(current_value, str):
            current_value = float(current_value.replace(',', ''))
        if isinstance(previous_value, str):
            previous_value = float(previous_value.replace(',', ''))

        if previous_value == 0:
            return 0.0 if current_value == 0 else 100.0
        return ((current_value - previous_value) / previous_value) * 100
    except (TypeError, ValueError):
        return 0.0

def calculate_inventory_turnover(inventory_data):
    try:
        cost_of_goods_sold = inventory_data.get('cost_of_goods_sold', 0)
        average_inventory = inventory_data.get('average_inventory_value', 0)

        if isinstance(cost_of_goods_sold, str):
            cost_of_goods_sold = float(cost_of_goods_sold.replace(',', ''))
        if isinstance(average_inventory, str):
            average_inventory = float(average_inventory.replace(',', ''))

        if average_inventory == 0:
            return 0

        turnover_ratio = cost_of_goods_sold / average_inventory
        return round(turnover_ratio, 2)
    except (TypeError, ValueError, ZeroDivisionError):
        return 0

def generate_advanced_analysis(report_data):
    """
    Generate sophisticated business analysis and insights based on report data.

    Args:
        report_data (dict): Comprehensive report data including financial and inventory metrics

    Returns:
        dict: Structured analysis including trends, risks, and opportunities
    """
    analysis = {
        'trends': [],
        'risks': [],
        'opportunities': [],
        'kpi_analysis': {}
    }

    try:
        total_revenue = parse_number(report_data['financial_overview']['total_revenue'])
        previous_revenue = parse_number(report_data['financial_overview'].get('previous_revenue', 0))

        formatted_revenue = format_number(total_revenue)
        formatted_previous_revenue = format_number(previous_revenue)

        revenue_growth = calculate_growth_rate(total_revenue, previous_revenue)

        if revenue_growth > 0:
            analysis['trends'].append({
                'metric': 'Revenue Growth',
                'value': revenue_growth,
                'impact': 'positive',
                'insight': f'Revenue grown by {revenue_growth:.1f}% (from {formatted_previous_revenue} to {formatted_revenue}) indicates strong market performance.'
            })

        try:
            net_profit = parse_number(report_data['financial_overview']['net_profit'])
            formatted_net_profit = format_number(net_profit)

            # Calculate profit margin using decimal values
            profit_margin = float(net_profit / total_revenue * 100) if float(total_revenue) > 0 else 0

            analysis['kpi_analysis']['profit_margin'] = {
                'value': profit_margin,
                'benchmark': 20.0,
                'status': 'above_target' if profit_margin > 20 else 'below_target',
                'formatted_profit': formatted_net_profit
            }
        except (ZeroDivisionError, KeyError, decimal.InvalidOperation) as e:
            logger.error(f"Error calculating profit margin: {str(e)}")
            analysis['kpi_analysis']['profit_margin'] = {
                'value': 0,
                'benchmark': 20.0,
                'status': 'below_target',
                'formatted_profit': format_number(0)
            }
            analysis['risks'].append({
                'category': 'Financial Data',
                'issue': 'Incomplete Profit Data',
                'impact': 'medium',
                'recommendation': 'Ensure complete financial data collection.'
            })

        stock_turnover = calculate_inventory_turnover(report_data['inventory_insights'])

        if stock_turnover > 0:
            if stock_turnover < 4:
                analysis['risks'].append({
                    'category': 'Inventory',
                    'issue': 'Low Stock Turnover',
                    'impact': 'medium',
                    'recommendation': 'Review procurement strategy and consider inventory optimization.'
                })
            elif stock_turnover > 12:
                analysis['opportunities'].append({
                    'category': 'Inventory',
                    'insight': 'High Stock Turnover',
                    'impact': 'positive',
                    'recommendation': 'Consider increasing stock levels to prevent stockouts.'
                })

    except Exception as e:
        logger.error(f"Error in advanced analysis generation: {str(e)}")
        # Ensure profit_margin key exists even in case of error
        analysis['kpi_analysis']['profit_margin'] = {
            'value': 0,
            'benchmark': 20.0,
            'status': 'below_target',
            'formatted_profit': format_number(0)
        }

    return analysis


def generate_pdf_report(report, start_date_str=None, end_date_str=None, report_data=None):
    if report_data is None:
        report_data = generate_comprehensive_report(report, start_date_str, end_date_str)

    # Initialize buffer and document with adjusted margins for better whitespace
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=60,
        leftMargin=60,
        topMargin=60,
        bottomMargin=40
    )

    # Enhanced style definitions with a more sophisticated color palette
    corporate_colors = {
        'primary': colors.HexColor('#1a237e'),     # Deep blue for primary elements
        'secondary': colors.HexColor('#303f9f'),   # Medium blue for secondary elements
        'accent': colors.HexColor('#7986cb'),      # Light blue for accents
        'background': colors.HexColor('#f5f6fa'),  # Light gray-blue for backgrounds
        'text': colors.HexColor('#2c3e50'),       # Dark gray-blue for main text
        'subtle': colors.HexColor('#90a4ae'),
        'warning': colors.HexColor('#fb8c00'),
        'chart_colors': [
            colors.HexColor('#1a237e'),
            colors.HexColor('#303f9f'),
            colors.HexColor('#7986cb'),
            colors.HexColor('#c5cae9'),
            colors.HexColor('#3949ab')
        ]
    }

    # Use the font in your styles
    styles = getSampleStyleSheet()
    
    # Main Title Style
    styles.add(ParagraphStyle(
        name='MainTitle',
        parent=styles['Title'],
        fontSize=28,
        spaceAfter=40,
        spaceBefore=30,
        alignment=TA_CENTER,
        textColor=corporate_colors['primary'],
        fontName='Helvetica-Bold',
        leading=34  # Increased leading for better readability
    ))

    # Section Header Style
    styles.add(ParagraphStyle(
        name='SectionHeader',
        parent=styles['Heading1'],
        fontSize=18,
        spaceBefore=25,
        spaceAfter=15,
        textColor=corporate_colors['secondary'],
        fontName='Helvetica-Bold',
        borderWidth=0,  # Removed border for cleaner look
        leading=22,
        alignment=TA_LEFT
    ))

    # Enhanced Sub-Header Style
    styles.add(ParagraphStyle(
        name='SubHeader',
        parent=styles['Heading2'],
        fontSize=14,
        spaceBefore=15,
        spaceAfter=10,
        textColor=corporate_colors['accent'],
        fontName='Helvetica-Bold',
        leading=18
    ))

    # Body Text Style
    if 'BodyText' not in styles:
        styles.add(ParagraphStyle(
            name='BodyText',
            parent=styles['Normal'],
            fontSize=11,
            textColor=corporate_colors['text'],
            spaceAfter=12,
            spaceBefore=12,
            leading=16,
            alignment=TA_LEFT
        ))

    # Style for insights and analysis text
    styles.add(ParagraphStyle(
        name='InsightText',
        parent=styles['BodyText'],
        fontSize=10,
        textColor=corporate_colors['secondary'],
        spaceBefore=6,
        spaceAfter=12,
        leading=14,
        leftIndent=20,
        bulletIndent=12,
        firstLineIndent=0
    ))

    # Enhanced Table Style
    enhanced_table_style = TableStyle([
        # Header styling
        ('BACKGROUND', (0, 0), (-1, 0), corporate_colors['background']),
        ('TEXTCOLOR', (0, 0), (-1, 0), corporate_colors['primary']),
        ('ALIGN', (0, 0), (-1, 0), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 15),
        ('TOPPADDING', (0, 0), (-1, 0), 15),
        
        # Body styling
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), corporate_colors['text']),
        ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 12),
        ('TOPPADDING', (0, 1), (-1, -1), 12),
        
        # Grid styling
        ('GRID', (0, 0), (-1, -1), 0.5, corporate_colors['subtle']),
        ('LINEBELOW', (0, 0), (-1, 0), 1, corporate_colors['primary']),
        
        # Alternate row colors for better readability
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, corporate_colors['background']])
    ])

    def prepare_revenue_chart_data(financial_overview):
        try:
            monthly_data = []
            cash_flow = financial_overview.get('monthly_cash_flow', [])
        
            for entry in cash_flow:
                if entry['transaction_type'] == 'income':
                    # Directly use the month field from the entry
                    transaction_date = entry['month']

                    revenue_amount = parse_number(entry['total_amount'])
                
                    monthly_data.append({
                        'month_year': transaction_date,
                        'month': transaction_date.strftime('%B'),
                        'year': transaction_date.year,
                        'revenue': revenue_amount
                    })

            # Sort the data by the datetime object
            monthly_data.sort(key=lambda x: x['month_year'])

            # Prepare the final response, keeping the year information
            formatted_data = [
                {
                    'month': data['month'],
                    'year': data['year'],
                    'revenue': data['revenue']
                }
                for data in monthly_data
            ]
        
            return {'monthly_revenue': formatted_data}
        
        except Exception as e:
            logger.error(f"Error preparing revenue chart data: {str(e)}")
            return {'monthly_revenue': []}

    def prepare_category_chart_data(income_breakdown):
        logger.info(f"Income breakdown data: {income_breakdown}")
    
        formatted_data = []
        total_amount = sum(parse_number(item['total_amount']) for item in income_breakdown)

        if total_amount > 0:
            for item in income_breakdown:
                amount = parse_number(item['total_amount'])
                if amount > 0:
                    category_name = item['category'] if item['category'] else 'Uncategorized Income'
                    formatted_data.append({
                        'category': category_name,
                        'total_amount': amount
                    })

        logger.info(f"Formatted category data: {formatted_data}")
        return formatted_data

    def prepare_inventory_chart_data(inventory_insights):
        try:
            products = inventory_insights['low_stock_products']
        
            # Handle case where products is already a list
            if isinstance(products, list):
                product_list = products[:10]  # Take first 10 items
            
                if not product_list:
                    logger.warning("No low stock products found")
                    return {
                        'product_categories': ['No Low Stock Items'],
                        'stock_levels': [[0]]
                    }

                # Ensure each product has required fields
                categories = []
                stock_levels = []
            
                for product in product_list:
                    name = product.get('name', 'Unknown Product')
                    stock = float(product.get('stock', 0))
                    categories.append(name)
                    stock_levels.append(stock)

                logger.info(f"Inventory chart data - Categories: {categories}, Levels: {stock_levels}")

                return {
                    'product_categories': categories,
                    'stock_levels': [stock_levels]  # Wrap in list to maintain expected format
                }
            else:
                logger.warning("Invalid product data format")
                return {
                    'product_categories': ['Data Unavailable'],
                    'stock_levels': [[0]]
                }
            
        except Exception as e:
            logger.error(f"Error in prepare_inventory_chart_data: {str(e)}")
            return {
                'product_categories': ['Data Unavailable'],
                'stock_levels': [[0]]
            }

    def validate_chart_dimensions(width, height, min_width=300, min_height=200):
        return max(width, min_width), max(height, min_height)

    formatted_category_data = prepare_category_chart_data(report_data['financial_overview']['income_breakdown'])
    logger.info(f"Category data being passed to chart: {formatted_category_data}")
    if not formatted_category_data:
        logger.warning("No valid category data available for pie chart")

    width, height = validate_chart_dimensions(400, 300)
    category_chart = create_category_pie_chart(formatted_category_data, width, height)

    content = []

    # Header Section
    period_text = f"Report Period: {start_date_str} to {end_date_str}" if start_date_str and end_date_str else "Complete Business Overview"
    content.append(Paragraph(f"Financial & Inventory Analysis Report", styles['MainTitle']))
    content.append(Paragraph(report.name, styles['SubHeader']))
    content.append(Paragraph(period_text, styles['BodyText']))
    content.append(Spacer(1, 30))

    # Executive Summary Section
    content.append(Paragraph("Executive Summary", styles['SectionHeader']))
    executive_summary = f"""
    This comprehensive analysis presents detailed insights into {report.name}'s financial performance,
    inventory status, and key business metrics for the period {period_text.lower()}. The report 
    highlights critical performance indicators and provides actionable insights for business optimization.
    """
    content.append(Paragraph(executive_summary.strip(), styles['BodyText']))
    content.append(Spacer(1, 25))

    # Executive Dashboard Section
    content.append(Paragraph("Executive Dashboard", styles['SectionHeader']))

    # Create KPI Summary Cards
    net_profit = parse_number(report_data['financial_overview']['net_profit'])
    total_revenue = parse_number(report_data['financial_overview']['total_revenue'])

    kpi_data = [
        ['Revenue', format_number(total_revenue, include_suffix=True),
         f"{calculate_growth_rate(report_data['financial_overview']['total_revenue'], report_data['financial_overview'].get('previous_revenue', 0)):+.1f}%"],
        ['Profit Margin', f"{((net_profit / total_revenue) * 100 if total_revenue else 0):.1f}%",
         'Target: 20%'],
        ['Customer Base', f"{report_data['performance_metrics']['customer_metrics']['total_customers']:,}",
         f"{calculate_growth_rate(report_data['performance_metrics']['customer_metrics']['total_customers'], report_data['performance_metrics']['customer_metrics'].get('previous_total_customers', 0)):+.1f}%"]
    ]


    # Create KPI cards with enhanced styling
    kpi_table = create_kpi_dashboard(kpi_data, doc.width, styles)
    content.append(kpi_table)
    content.append(Spacer(1, 30))

    # Revenue Trend Visualization
    content.append(Paragraph("Revenue Performance Trend", styles['SubHeader']))
    width, height = validate_chart_dimensions(500, 200)
    revenue_chart = create_revenue_trend_chart(
        prepare_revenue_chart_data(report_data['financial_overview']), 
        width, 
        height
    )
    content.append(revenue_chart)
    content.append(Spacer(1, 20))

    # Category Distribution
    content.append(Paragraph("Revenue Distribution by Category", styles['SubHeader']))
    width, height = validate_chart_dimensions(400, 300)
    category_chart = create_category_pie_chart(
        prepare_category_chart_data(report_data['financial_overview']['income_breakdown']), 
        width, 
        height
    )
    content.append(category_chart)

    # Add page break before detailed analysis
    content.append(PageBreak())

    # Detailed Analysis Sections
    advanced_analysis = generate_advanced_analysis(report_data)

    # Performance Insights Section
    content.append(Paragraph("Key Performance Insights", styles['SectionHeader']))

    for trend in advanced_analysis['trends']:
        content.append(Paragraph(
            f"• {trend['insight']}",
            styles['InsightText']
        ))

    # Risk Assessment Section
    content.append(Paragraph("Risk Assessment", styles['SectionHeader']))

    for risk in advanced_analysis['risks']:
        content.append(Paragraph(
            f"• {risk['category']}: {risk['issue']} - {risk['recommendation']}",
            styles['InsightText']
        ))

    # Financial Performance Section with enhanced table structure
    financial_data = report_data['financial_overview']

    total_revenue = parse_number(financial_data['total_revenue'])
    cost_of_services = parse_number(financial_data['cost_of_services'])
    operating_expenses = parse_number(financial_data['operating_expenses'])
    net_profit = parse_number(financial_data['net_profit'])

    financial_summary = [
        ['Key Metrics', 'Amount', 'Analysis'],
        ['Total Revenue',
         format_number(total_revenue, include_suffix=True),
         'Total business income generated during the period'],
        ['Cost of Services',
         format_number(cost_of_services, include_suffix=True),
         'Direct costs associated with service delivery'],
        ['Operating Expenses',
         format_number(operating_expenses, include_suffix=True),
         'General and administrative expenses'],
        ['Net Profit',
         format_number(net_profit, include_suffix=True),
         'Final profit after all deductions']
    ]

    # Create table with adjusted column widths for better proportions
    financial_table = Table(financial_summary, colWidths=[doc.width * 0.25, doc.width * 0.25, doc.width * 0.5])
    financial_table.setStyle(enhanced_table_style)
    content.append(financial_table)
    content.append(Spacer(1, 25))

    content.append(Paragraph("Revenue by Category Analysis", styles['SectionHeader']))
    
    revenue_insight = """
    The following analysis breaks down revenue streams by category, highlighting key contributors 
    to overall business performance. This segmentation provides insights into revenue distribution 
    and helps identify areas for potential growth or optimization.
    """
    content.append(Paragraph(revenue_insight.strip(), styles['BodyText']))
    
    income_data = report_data['financial_overview']['income_breakdown']
    
    # Create revenue breakdown table with enhanced formatting
    income_table_data = [[
        'Revenue Category',
        'Amount',
        'Transaction Volume',
        'Revenue Share',
        'Trend'
    ]]
    
    total_revenue = parse_number(report_data['financial_overview']['total_revenue'])

    for category in income_data:
        amount = parse_number(category['total_amount'])
        share = (amount / total_revenue) * 100 if total_revenue > 0 else 0
        trend_indicator = '↑' if category.get('growth_rate', 0) > 0 else '↓'
        trend_color = corporate_colors['success'] if category.get('growth_rate', 0) > 0 else corporate_colors['warning']

        income_table_data.append([
            category['category'],
            format_number(amount, include_suffix=True),
            f"{category['transaction_count']:,}",
            f"{share:.1f}%",
            Paragraph(f"{trend_indicator} {abs(category.get('growth_rate', 0)):.1f}%",
                     ParagraphStyle('TrendStyle',
                                  parent=styles['BodyText'],
                                  textColor=trend_color))
        ])

    income_table = Table(income_table_data,
                        colWidths=[doc.width * 0.25, doc.width * 0.2,
                                 doc.width * 0.25, doc.width * 0.2,
                                 doc.width * 0.15])
    income_table.setStyle(enhanced_table_style)
    content.append(income_table)
    content.append(Spacer(1, 20))

    # Inventory Analytics Section with Enhanced Metrics
    content.append(Paragraph("Inventory Analytics & Management", styles['SectionHeader']))
    
    inventory_insight = """
    This section provides a comprehensive overview of inventory status, movement patterns, and key 
    performance indicators. The analysis helps identify potential stock optimization opportunities 
    and areas requiring attention.
    """
    content.append(Paragraph(inventory_insight.strip(), styles['BodyText']))
    
    inventory_data = report_data['inventory_insights']
    
    # Handle low stock products count properly
    low_stock_count = len(inventory_data['low_stock_products']) if isinstance(inventory_data['low_stock_products'], list) else 0
    
    # Parse total stock value using the currency formatter
    total_stock_value = parse_number(inventory_data['total_stock_value'])

    # Create enhanced inventory metrics table with proper data handling
    inventory_metrics = [
        ['Metric', 'Current Value', 'Status', 'Recommendation'],
        ['Total SKUs',
         str(inventory_data['total_product_count']),
         'Active Inventory',
         'Monitor product mix diversity'],
        ['Total Stock Value',
         format_number(total_stock_value, include_suffix=True),
         'Invested Capital',
         'Optimize working capital allocation'],
        ['Low Stock Items',
         str(low_stock_count),
         'Requires Attention' if low_stock_count > 0 else 'Optimal',
         'Review reorder points and lead times'],
        ['Inventory Turnover',
         f"{float(inventory_data.get('inventory_turnover', 0)):.2f}x",
         'Stock Efficiency',
         'Analyze slow-moving items']
    ]

    inventory_table = Table(
        inventory_metrics,
        colWidths=[doc.width * 0.2, doc.width * 0.2, doc.width * 0.2, doc.width * 0.4]
    )
    inventory_table.setStyle(enhanced_table_style)
    content.append(inventory_table)
    content.append(Spacer(1, 20))

    # Enhanced inventory chart handling with better error checking
    try:
        content.append(Paragraph("Inventory Stock Levels", styles['SubHeader']))
        inventory_chart_data = prepare_inventory_chart_data(report_data['inventory_insights'])
        
        logger.info(f"Prepared inventory data: {inventory_chart_data}")

        if (isinstance(inventory_chart_data['product_categories'], list) and 
            inventory_chart_data['product_categories'][0] not in ['No Low Stock Items', 'Data Unavailable']):
            width, height = validate_chart_dimensions(500, 200)
            inventory_chart = create_inventory_bar_chart(inventory_chart_data, width, height)
            content.append(inventory_chart)
        else:
            content.append(Paragraph("No low stock items to display", styles['BodyText']))

        content.append(Spacer(1, 20))
    except Exception as e:
        logger.error(f"Error adding inventory chart to report: {str(e)}")
        content.append(Paragraph("Unable to generate inventory chart", styles['BodyText']))
        content.append(Spacer(1, 20))

    # Business Performance Metrics section with enhanced currency handling
    content.append(Paragraph("Business Performance & Customer Metrics", styles['SectionHeader']))

    performance_insight = """
    This section analyzes key business performance indicators and customer metrics, providing insights
    into operational efficiency and customer engagement levels. The metrics help identify trends and
    areas for strategic focus.
    """
    content.append(Paragraph(performance_insight.strip(), styles['BodyText']))

    performance_data = report_data['performance_metrics']
    
    # Parse currency values for average order calculations
    current_aov = parse_number(performance_data['order_performance']['average_order_value'])
    previous_aov = parse_number(performance_data['order_performance'].get('previous_average_order_value', 0))

    performance_metrics = [
        ['Key Performance Indicator', 'Current Value', 'Previous Period', 'Change'],
        ['Total Customers',
         str(performance_data['customer_metrics']['total_customers']),
         str(performance_data['customer_metrics'].get('previous_total_customers', 0)),
         calculate_change(performance_data['customer_metrics']['total_customers'],
                        performance_data['customer_metrics'].get('previous_total_customers', 0))],
        ['Average Order Value',
         format_number(current_aov, include_suffix=True),
         format_number(previous_aov, include_suffix=True),
         calculate_change(current_aov, previous_aov)],
        ['Customer Retention Rate',
         f"{float(performance_data['customer_metrics'].get('retention_rate', 0)):.1f}%",
         f"{float(performance_data['customer_metrics'].get('previous_retention_rate', 0)):.1f}%",
         calculate_change(float(performance_data['customer_metrics'].get('retention_rate', 0)),
                        float(performance_data['customer_metrics'].get('previous_retention_rate', 0)))],
        ['Average Orders per Customer',
         f"{float(performance_data['customer_metrics']['average_orders_per_customer']):.2f}",
         f"{float(performance_data['customer_metrics'].get('previous_average_orders', 0)):.2f}",
         calculate_change(
             float(performance_data['customer_metrics']['average_orders_per_customer']),
             float(performance_data['customer_metrics'].get('previous_average_orders', 0)))]
    ]
    
    performance_table = Table(performance_metrics, colWidths=[doc.width * 0.4, doc.width * 0.2,
                                                            doc.width * 0.2, doc.width * 0.2])
    performance_table.setStyle(enhanced_table_style)
    content.append(performance_table)
    content.append(Spacer(1, 20))

    # Recommendations Section
    content.append(Paragraph("Strategic Action Items", styles['SectionHeader']))
    recommendations = generate_strategic_recommendations(advanced_analysis)
    
    for rec in recommendations:
        content.append(Paragraph(
            f"• {rec['action']}: {rec['rationale']}",
            styles['InsightText']
        ))

    # Enhanced footer with page numbers
    def add_page_number(canvas, doc):
        page_num = canvas.getPageNumber()
        text = f"Page {page_num}"
        canvas.saveState()
        canvas.setFont('Helvetica', 9)
        canvas.setFillColor(corporate_colors['subtle'])
        canvas.drawRightString(doc.pagesize[0] - 60, 30, text)
        canvas.drawString(60, 30, report.name)
        canvas.line(60, 40, doc.pagesize[0] - 60, 40)
        canvas.restoreState()

    # Build document
    doc.build(content, onFirstPage=add_page_number, onLaterPages=add_page_number)

    # Finalize PDF
    pdf_content = buffer.getvalue()
    buffer.close()

    formatted_date = datetime.now().strftime("%Y%m%d")
    pdf_file = ContentFile(pdf_content)
    pdf_file.name = f"{report.name}_Financial_Analysis_{formatted_date}.pdf"

    return pdf_file

def calculate_change(current, previous):
    """Calculate percentage change with appropriate formatting"""
    if previous == 0:
        return "N/A"
    change = ((current - previous) / previous) * 100
    return f"{'+' if change > 0 else ''}{change:.1f}%"

def create_kpi_dashboard(kpi_data, width, styles):
    """
    Create an enhanced KPI dashboard with visual indicators
    
    Args:
        kpi_data (list): List of KPI data to display
        width (float): Width of the table
        styles (dict): ReportLab stylesheet containing base styles
        
    Returns:
        Table: A formatted ReportLab Table object containing the KPI dashboard
    """
    table_data = []
    
    # Create a custom KPI cell style
    kpi_cell_style = ParagraphStyle(
        'KPICell',
        parent=styles['BodyText'],
        fontSize=12,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#2c3e50')
    )
    
    for kpi in kpi_data:
        formatted_cells = [
            Paragraph(str(kpi[0]), kpi_cell_style),
            Paragraph(str(kpi[1]), kpi_cell_style),
            Paragraph(str(kpi[2]), kpi_cell_style)
        ]
        table_data.append(formatted_cells)
    
    # Create and style the table
    table = Table(table_data, colWidths=[width/3.0]*3)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f5f6fa')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#2c3e50')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 12),
        ('TOPPADDING', (0, 0), (-1, -1), 20),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 20),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e8eaf6')),
        ('ROUNDEDCORNERS', (0, 0), (-1, -1), 8)
    ]))
    
    return table

def generate_strategic_recommendations(analysis):
    recommendations = []

    # Safely access profit margin data
    profit_margin_data = analysis.get('kpi_analysis', {}).get('profit_margin', {})
    if profit_margin_data.get('status') == 'below_target':
        recommendations.append({
            'action': 'Implement Cost Optimization Program',
            'rationale': 'Current profit margins are below target. Focus on operational efficiency and pricing strategy.',
            'priority': 'high'
        })

    # Add default recommendation if list is empty
    if not recommendations:
        recommendations.append({
            'action': 'Review Business Metrics',
            'rationale': 'Conduct comprehensive review of all business metrics to identify areas for improvement.',
            'priority': 'medium'
        })

    return recommendations 

def send_report_email(report, recipient_email, request=None, include_summary=True, include_charts=True, start_date_str=None, end_date_str=None):
    try:
        logger.info(f"Starting email preparation for report: {report.name}")
        
        # Validate and process date range
        if not start_date_str:
            start_date_str = report.start_date.strftime('%Y-%m-%d') if report.start_date else report.created_at.strftime('%Y-%m-%d')
        if not end_date_str:
            end_date_str = report.end_date.strftime('%Y-%m-%d') if report.end_date else timezone.now().strftime('%Y-%m-%d')
            
        logger.info(f"Generating report for date range: {start_date_str} to {end_date_str}")
        
        # Generate report with specified date range
        report_data = generate_comprehensive_report(report, start_date_str, end_date_str)

        logger.info(f"Generated report data for period: {start_date_str} to {end_date_str}")
        logger.info(f"Financial Overview - Total Revenue: ${report_data['financial_overview']['total_revenue']:,.2f}")

        pdf_file = generate_pdf_report(report, report_data=report_data)
        pdf_content = pdf_file.read()

        # Update email subject to include date range
        period_str = f"{start_date_str} to {end_date_str}"
        email_subject = f"Financial Report: {report.name} ({period_str})"

        email_template = Report.objects.filter(
            is_template=True,
            name="Basic Email Template"
        ).first()

        if not email_template:
            raise ValueError("Email template not found")

        msg = EmailMultiAlternatives(
            subject=email_subject,
            body=body_template.render(context),
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[recipient_email]
        )

        msg.attach(
            filename=f"{report.name}_report_{start_date_str}_to_{end_date_str}.pdf",
            content=pdf_content,
            mimetype='application/pdf'
        )

        logger.info(f"Sending email with PDF for period {period_str}")
        msg.send(fail_silently=False)
        
        logger.info(f"Report email sent successfully to {recipient_email}")
        return True

    except Exception as e:
        logger.error(f"Error sending report email: {str(e)}")
        raise EmailSendingError(str(e))

def export_styled_report(report, start_date_str=None, end_date_str=None):
    """
    Generates a visually enhanced financial report with professional styling and formatting.
    Returns a single CSV file with enhanced formatting.

    Args:
        report: Report object containing the business data
        start_date_str: Optional start date for filtering (str)
        end_date_str: Optional end date for filtering (str)

    Returns:
        ContentFile: CSV file with enhanced formatting
    """
    # Define formatting constants
    SEPARATOR_LINE = ["=" * 80]
    SECTION_SEPARATOR = ["-" * 80]
    INDENT = "    "

    report_data = generate_comprehensive_report(report, start_date_str, end_date_str)
    rows = []

    # Header Section with clear separation
    rows.extend(SEPARATOR_LINE)
    rows.extend([
        ["EXECUTIVE SUMMARY"],
        [""],
        [f"{INDENT}Report Name:", report.name],
        [f"{INDENT}Analysis Period:", f"{start_date_str} to {end_date_str}" if start_date_str and end_date_str else "Complete Business Overview"],
        [f"{INDENT}Generated:", datetime.now().strftime('%B %d, %Y %H:%M')],
        [f"{INDENT}Author:", str(report.created_by)],
        [f"{INDENT}Last Modified:", report.updated_at.strftime('%B %d, %Y %H:%M')],
        [""]
    ])

    # Financial Highlights Section
    rows.extend(SECTION_SEPARATOR)
    fo = report_data.get('financial_overview', {})
    
    # Use parse_number to ensure numeric values
    total_revenue = parse_number(fo.get('total_revenue', 0))
    cost_of_services = parse_number(fo.get('cost_of_services', 0))
    operating_expenses = parse_number(fo.get('operating_expenses', 0))
    
    calculated_gross_margin = total_revenue - cost_of_services
    net_profit = parse_number(fo.get('net_profit', 0)) or (calculated_gross_margin - operating_expenses)

    rows.extend([
        ["FINANCIAL HIGHLIGHTS"],
        [""],
        [f"{INDENT}Key Performance Indicators", "Amount", "% of Revenue", "Analysis"],
        [f"{INDENT}Total Revenue", format_number(total_revenue), "100.00%", "Primary income stream"],
        [f"{INDENT}Cost of Services", format_number(cost_of_services),
         f"{(cost_of_services/total_revenue*100 if total_revenue else 0):.2f}%", "Direct service costs"],
        [f"{INDENT}Gross Margin", format_number(calculated_gross_margin),
         f"{(calculated_gross_margin/total_revenue*100 if total_revenue else 0):.2f}%", "Operating efficiency"],
        [f"{INDENT}Operating Expenses", format_number(operating_expenses),
         f"{(operating_expenses/total_revenue*100 if total_revenue else 0):.2f}%", "Overhead costs"],
        [f"{INDENT}Net Profit", format_number(net_profit),
         f"{(net_profit/total_revenue*100 if total_revenue else 0):.2f}%", "Bottom line"],
        [""]
    ])

    # Revenue Analysis Section
    rows.extend(SECTION_SEPARATOR)
    rows.extend([
        ["REVENUE ANALYSIS"],
        [""],
        [f"{INDENT}Revenue Stream", "Amount", "Transaction Volume", "Share of Revenue"]
    ])

    for category in fo.get('income_breakdown', []):
        transaction_count = category.get('transaction_count', 0)
        total_amount = parse_number(category.get('total_amount', 0))
        revenue_share = (total_amount / total_revenue) if total_revenue > 0 else 0

        rows.append([
            f"{INDENT}{category.get('category', 'Uncategorized')}",
            format_number(total_amount),
            f"{transaction_count:,}",
            f"{revenue_share:.2%}"
        ])
    rows.append([""])

    # Operational Metrics Section
    rows.extend(SECTION_SEPARATOR)
    inventory_insights = report_data.get('inventory_insights', {})
    pm = report_data.get('performance_metrics', {})

    rows.extend([
        ["OPERATIONAL METRICS"],
        [""],
        [f"{INDENT}Category", "Current Value", "Target", "Impact"],
        [""],
        [f"{INDENT}Inventory Management"],
        [f"{INDENT}{INDENT}Total Stock Value", format_number(parse_number(inventory_insights.get('total_stock_value', 0))),
         "Variable", "Working Capital"],
        [f"{INDENT}{INDENT}Low Stock Items", str(len(inventory_insights.get('low_stock_products', []))),
         "0", "Service Level"],
        [""],
        [f"{INDENT}Customer Metrics"],
        [f"{INDENT}{INDENT}Active Customers", f"{pm.get('customer_metrics', {}).get('total_customers', 0):,}",
         "Growing", "Market Share"],
        [f"{INDENT}{INDENT}Avg. Order Value", format_number(parse_number(pm.get('order_performance', {}).get('average_order_value', 0))),
         "Growing", "Revenue Growth"]
    ])

    rows.extend(SEPARATOR_LINE)

    # Write to CSV with enhanced formatting
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    for row in rows:
        writer.writerow(row if isinstance(row, list) else [row])

    csv_content = buffer.getvalue().encode('utf-8')
    buffer.close()

    # Create file with formatted name
    csv_file = ContentFile(csv_content)
    formatted_date = datetime.now().strftime("%Y%m%d_%H%M")
    csv_file.name = f"{report.name}_financial_analysis_{formatted_date}.csv"

    return csv_file

def create_excel_styles(workbook):
    """Create and return a dictionary of sophisticated named styles for professional Excel formatting."""
    # Enhanced color palette using professional, accessible colors
    colors = {
        'primary': '1E3A8A',      # Rich navy blue for headers
        'secondary': 'F1F5F9',    # Soft blue-gray for alternate rows
        'accent1': '2563EB',      # Vivid blue for highlights
        'accent2': 'F8FAFC',      # Lightest blue-gray for subtle emphasis
        'border': 'CBD5E1',       # Medium gray for borders
        'positive': '059669',     # Emerald green for positive values
        'negative': 'DC2626',     # Crimson red for negative values
        'neutral': '475569',      # Slate gray for neutral text
        'muted': '94A3B8'         # Muted blue-gray for secondary text
    }

    # Sophisticated font configurations
    header_font = Font(
        name='Segoe UI',
        size=12,
        bold=True,
        color=colors['primary']
    )
    
    subheader_font = Font(
        name='Segoe UI',
        size=11,
        bold=True,
        color=colors['accent1']
    )
    
    normal_font = Font(
        name='Segoe UI',
        size=10,
        color=colors['neutral']
    )

    # Enhanced border styles
    thin_border = Side(style='thin', color=colors['border'])
    medium_border = Side(style='medium', color=colors['primary'])
    
    standard_border = Border(
        left=thin_border,
        right=thin_border,
        top=thin_border,
        bottom=thin_border
    )
    
    header_border = Border(
        left=thin_border,
        right=thin_border,
        top=medium_border,
        bottom=medium_border
    )

    # Enhanced header style
    header_style = NamedStyle(name='header_style')
    header_style.font = header_font
    header_style.fill = PatternFill(
        start_color=colors['secondary'],
        end_color=colors['secondary'],
        fill_type='solid'
    )
    header_style.alignment = Alignment(
        horizontal='left',
        vertical='center',
        wrap_text=True
    )
    header_style.border = header_border
    workbook.add_named_style(header_style)

    # Enhanced subheader style
    subheader_style = NamedStyle(name='subheader_style')
    subheader_style.font = subheader_font
    subheader_style.fill = PatternFill(
        start_color=colors['accent2'],
        end_color=colors['accent2'],
        fill_type='solid'
    )
    subheader_style.alignment = Alignment(
        horizontal='left',
        vertical='center'
    )
    subheader_style.border = standard_border
    workbook.add_named_style(subheader_style)

    # Sophisticated currency style
    currency_style = NamedStyle(
        name='currency_style',
        number_format='_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
    )
    currency_style.font = normal_font
    currency_style.alignment = Alignment(horizontal='right')
    currency_style.border = standard_border
    workbook.add_named_style(currency_style)

    # Enhanced currency negative style
    currency_negative_style = NamedStyle(
        name='currency_negative_style',
        number_format='[Red]_($* #,##0.00_);[Red]_($* (#,##0.00);_($* "-"??_);_(@_)'
    )
    currency_negative_style.font = Font(
        name='Segoe UI',
        size=10,
        color=colors['negative']
    )
    currency_negative_style.alignment = Alignment(horizontal='right')
    currency_negative_style.border = standard_border
    workbook.add_named_style(currency_negative_style)

    # Enhanced percentage style
    percentage_style = NamedStyle(
        name='percentage_style',
        number_format='0.00%'
    )
    percentage_style.font = normal_font
    percentage_style.alignment = Alignment(horizontal='center')
    percentage_style.border = standard_border
    workbook.add_named_style(percentage_style)

    # Enhanced metric style for KPIs
    metric_style = NamedStyle(name='metric_style')
    metric_style.font = Font(
        name='Segoe UI',
        size=11,
        bold=True,
        color=colors['accent1']
    )
    metric_style.alignment = Alignment(
        horizontal='center',
        vertical='center'
    )
    metric_style.border = standard_border
    workbook.add_named_style(metric_style)

    # Enhanced normal style
    normal_style = NamedStyle(name='normal_style')
    normal_style.font = normal_font
    normal_style.alignment = Alignment(
        horizontal='left',
        vertical='center'
    )
    normal_style.border = standard_border
    workbook.add_named_style(normal_style)

    return {
        'header': header_style,
        'subheader': subheader_style,
        'currency': currency_style,
        'currency_negative': currency_negative_style,
        'percentage': percentage_style,
        'normal': normal_style,
        'metric': metric_style
    }

def format_worksheet(worksheet, data_frame, styles, start_row=1, alternate_rows=True):
    """Apply sophisticated formatting to worksheet with enhanced visual appeal."""
    # Enhanced print setup
    worksheet.page_setup.orientation = worksheet.ORIENTATION_LANDSCAPE
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4
    worksheet.print_options.gridLines = False
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True

    # Ensure sheet visibility
    worksheet.sheet_state = 'visible'

    # Freeze panes and set zoom
    worksheet.freeze_panes = f'A{start_row + 1}'
    worksheet.sheet_view.zoomScale = 100

    # Intelligent column width adjustment
    for idx, col in enumerate(data_frame.columns, 1):
        col_letter = get_column_letter(idx)
        max_length = max(
            data_frame[col].astype(str).apply(len).max(),
            len(str(col))
        )
        
        # Sophisticated width calculation based on content type
        if any(text in col.lower() for text in ['notes', 'description', 'analysis']):
            worksheet.column_dimensions[col_letter].width = min(max_length * 0.85, 50)
        elif any(text in col.lower() for text in ['date', 'period']):
            worksheet.column_dimensions[col_letter].width = max(len(str(col)) * 1.2, 12)
        elif any(text in col.lower() for text in ['amount', 'revenue', 'cost', 'value']):
            worksheet.column_dimensions[col_letter].width = max(len(str(col)) * 1.3, 15)
        else:
            worksheet.column_dimensions[col_letter].width = min(max_length * 1.1, 25)

    # Apply enhanced header styling
    header_row = worksheet[start_row]
    for cell in header_row:
        cell.style = styles['header']

    # Apply sophisticated data styling with improved alternate rows
    for idx, col in enumerate(data_frame.columns, 1):
        col_letter = get_column_letter(idx)
        
        for row in range(start_row + 1, worksheet.max_row + 1):
            cell = worksheet[f"{col_letter}{row}"]
            
            # Enhanced style selection based on content
            if any(text in col.lower() for text in ['amount', 'revenue', 'cost', 'value']):
                try:
                    value = float(cell.value or 0)
                    cell.style = styles['currency_negative'] if value < 0 else styles['currency']
                except (ValueError, TypeError):
                    cell.style = styles['normal']
            
            elif any(text in col.lower() for text in ['percentage', 'share', 'ratio', 'margin']):
                cell.style = styles['percentage']
                
            elif any(text in col.lower() for text in ['metric', 'kpi', 'indicator']):
                cell.style = styles['metric']
            
            else:
                cell.style = styles['normal']

            # Enhanced alternate row styling
            if alternate_rows and row % 2 == 0:
                cell.fill = PatternFill(
                    start_color='F8FAFC',
                    end_color='F8FAFC',
                    fill_type='solid'
                )

    # Add enhanced total row
    if any(col.lower() in ['amount', 'revenue', 'cost', 'value'] for col in data_frame.columns):
        total_row = worksheet.max_row + 1
        
        # Style the total row header
        total_cell = worksheet.cell(row=total_row, column=1, value='Total')
        total_cell.style = styles['subheader']
        
        # Add subtotals for numeric columns
        for idx, col in enumerate(data_frame.columns, 1):
            if col.lower() in ['amount', 'revenue', 'cost', 'value']:
                cell = worksheet.cell(row=total_row, column=idx)
                cell.value = f'=SUM({get_column_letter(idx)}{start_row + 1}:{get_column_letter(idx)}{total_row-1})'
                cell.style = styles['currency']
                
                # Add bottom border to total row
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='double')
                )

    # Add auto-filter to header row
    worksheet.auto_filter.ref = f"A{start_row}:{get_column_letter(worksheet.max_column)}{worksheet.max_row}"

def export_report_to_excel(report, start_date_str=None, end_date_str=None):
    """
    Generate a consolidated Excel report with professional formatting and detailed analysis.

    Args:
        report: Report instance containing business data
        start_date_str: Optional start date in YYYY-MM-DD format
        end_date_str: Optional end date in YYYY-MM-DD format

    Returns:
        ContentFile: Professionally formatted Excel file content
    """
    try:
        report_data = generate_comprehensive_report(report, start_date_str, end_date_str)
        buffer = io.BytesIO()

        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            workbook = writer.book
            styles = create_excel_styles(workbook)

            # Set professional workbook properties
            workbook.properties.title = f"{report.name} - Financial Analysis"
            workbook.properties.subject = "Business Performance Report"
            workbook.properties.creator = str(report.created_by)
            workbook.properties.created = datetime.now()
            workbook.properties.category = "Financial Reports"

            # Create consolidated DataFrame
            data_frames = []
            current_row = 0

            # 1. Report Header
            header_data = pd.DataFrame([
                {'Section': 'REPORT INFORMATION', 'Metric': '', 'Value': '', 'Description': ''},
                {'Section': '', 'Metric': 'Report Name', 'Value': report.name, 'Description': 'Comprehensive Business Analysis'},
                {'Section': '', 'Metric': 'Analysis Period', 
                 'Value': f"{start_date_str} to {end_date_str}" if start_date_str and end_date_str else "Complete Business Overview",
                 'Description': 'Analysis Timeframe'},
                {'Section': '', 'Metric': 'Generated Date', 
                 'Value': timezone.now().strftime('%Y-%m-%d %H:%M:%S'),
                 'Description': 'Report Creation Timestamp'},
                {'Section': '', 'Metric': 'Report Author', 
                 'Value': str(report.created_by),
                 'Description': 'Report Owner'},
                {'Section': '', 'Metric': '', 'Value': '', 'Description': ''}
            ])
            data_frames.append(header_data)
            current_row += len(header_data)

            # 2. Financial Overview
            fo = report_data.get('financial_overview', {})
            total_revenue = parse_number(fo.get('total_revenue', 0))
            cost_of_services = parse_number(fo.get('cost_of_services', 0))
            operating_expenses = parse_number(fo.get('operating_expenses', 0))
            calculated_gross_margin = total_revenue - cost_of_services
            net_profit = parse_number(fo.get('net_profit', calculated_gross_margin - operating_expenses))

            financial_data = pd.DataFrame([
                {'Section': 'FINANCIAL OVERVIEW', 'Metric': '', 'Value': '', 'Description': ''},
                {'Section': '', 'Metric': 'Total Revenue', 'Value': total_revenue, 
                 'Description': 'Gross business income across all revenue streams'},
                {'Section': '', 'Metric': 'Cost of Services', 'Value': cost_of_services,
                 'Description': 'Direct costs associated with service delivery'},
                {'Section': '', 'Metric': 'Gross Margin', 'Value': calculated_gross_margin,
                 'Description': 'Revenue less cost of services'},
                {'Section': '', 'Metric': 'Operating Expenses', 'Value': operating_expenses,
                 'Description': 'Operational and administrative costs'},
                {'Section': '', 'Metric': 'Net Profit', 'Value': net_profit,
                 'Description': 'Final profit after all deductions'},
                {'Section': '', 'Metric': '', 'Value': '', 'Description': ''}
            ])
            data_frames.append(financial_data)
            current_row += len(financial_data)

            # 3. Revenue Analysis
            revenue_header = pd.DataFrame([
                {'Section': 'REVENUE ANALYSIS', 'Metric': '', 'Value': '', 'Description': ''}
            ])
            data_frames.append(revenue_header)
            current_row += 1

            income_breakdown = fo.get('income_breakdown', [])
            if income_breakdown:
                revenue_data = pd.DataFrame([
                    {'Section': '',
                     'Metric': cat.get('category', 'Uncategorized'),
                     'Value': parse_number(cat.get('total_amount', 0)),
                     'Description': f"Count: {cat.get('transaction_count', 0)} | "
                                  f"Avg Value: {format_number(parse_number(cat.get('total_amount', 0)) / cat.get('transaction_count', 1) if cat.get('transaction_count', 0) > 0 else 0)} | "
                                  f"Share: {(parse_number(cat.get('total_amount', 0)) / total_revenue * 100 if total_revenue > 0 else 0):.1f}%"
                    } for cat in income_breakdown
                ])
                data_frames.append(revenue_data)
                current_row += len(revenue_data)

            # 4. Inventory Insights
            inventory_insights = report_data.get('inventory_insights', {})
            low_stock_count = len(inventory_insights.get('low_stock_products', [])) if isinstance(inventory_insights.get('low_stock_products', []), list) else 0

            inventory_data = pd.DataFrame([
                {'Section': '', 'Metric': '', 'Value': '', 'Description': ''},
                {'Section': 'INVENTORY MANAGEMENT', 'Metric': '', 'Value': '', 'Description': ''},
                {'Section': '', 'Metric': 'Total Products',
                 'Value': inventory_insights.get('total_product_count', 0),
                 'Description': 'Current inventory item count'},
                {'Section': '', 'Metric': 'Total Stock Value',
                 'Value': parse_number(inventory_insights.get('total_stock_value', 0)),
                 'Description': 'Current inventory value'},
                {'Section': '', 'Metric': 'Low Stock Items',
                 'Value': low_stock_count,
                 'Description': 'Items below threshold'},
                {'Section': '', 'Metric': '', 'Value': '', 'Description': ''}
            ])

            data_frames.append(inventory_data)
            current_row += len(inventory_data)

            # 5. Performance Metrics
            pm = report_data.get('performance_metrics', {})
            customer_metrics = pm.get('customer_metrics', {})
            order_performance = pm.get('order_performance', {})

            performance_data = pd.DataFrame([
                {'Section': 'PERFORMANCE METRICS', 'Metric': '', 'Value': '', 'Description': ''},
                {'Section': '', 'Metric': 'Customer Base',
                 'Value': customer_metrics.get('total_customers', 0),
                 'Description': 'Total active customers'},
                {'Section': '', 'Metric': 'Average Orders/Customer',
                 'Value': float(customer_metrics.get('average_orders_per_customer', 0)),
                 'Description': 'Customer engagement level'},
                {'Section': '', 'Metric': 'Average Order Value',
                 'Value': parse_number(order_performance.get('average_order_value', 0)),
                 'Description': 'Transaction value analysis'}
            ])

            data_frames.append(performance_data)

            # Combine all sections
            consolidated_data = pd.concat(data_frames, ignore_index=True)
            workbook.create_sheet('Business Analysis')

            consolidated_data.to_excel(writer, sheet_name='Business Analysis', index=False)

            # Format the consolidated worksheet
            worksheet = writer.sheets['Business Analysis']
            
            # Apply formatting
            for row in range(1, len(consolidated_data) + 2):  # +2 for header row
                for col in range(1, 5):  # 4 columns
                    cell = worksheet.cell(row=row, column=col)
                    
                    # Section headers
                    if consolidated_data.iloc[row-2]['Section'] != '':
                        cell.style = styles['header']
                    # Regular rows
                    else:
                        cell.style = styles['normal']
                        
                    # Format value column for numbers
                    if col == 3 and isinstance(consolidated_data.iloc[row-2]['Value'], (int, float)):
                        cell.number_format = '#,##0.00'

            # Adjust column widths
            worksheet.column_dimensions['A'].width = 20  # Section
            worksheet.column_dimensions['B'].width = 25  # Metric
            worksheet.column_dimensions['C'].width = 15  # Value
            worksheet.column_dimensions['D'].width = 50  # Description

        excel_content = buffer.getvalue()
        buffer.close()

        formatted_date = datetime.now().strftime("%Y%m%d_%H%M")
        excel_file = ContentFile(excel_content)
        excel_file.name = f"{report.name}_Financial_Analysis_{formatted_date}.xlsx"

        return excel_file

    except Exception as e:
        logger.error(f"Error generating Excel report: {str(e)}", exc_info=True)
        logger.error(f"Report data structure: {report_data.keys() if 'report_data' in locals() else 'No report data available'}")
        raise Exception(f"Failed to generate Excel report: {str(e)}")
        

def save_generated_file(report, file_content, file_type):
    """
    Save generated report file to ReportFile model.

    Args:
        report (Report): The report object
        file_content (ContentFile): Generated file content
        file_type (str): Type of file (pdf, excel, csv)

    Returns:
        ReportFile: Created file instance
    """
    try:
        # Create a default entry if no entries exist
        entry, _ = ReportEntry.objects.get_or_create(
            report=report,
            title=f"Generated {file_type.upper()} Report",
            defaults={
                'content': f"Automatically generated {file_type.upper()} report",
                'created_by': report.created_by,
                'last_modified_by': report.created_by
            }
        )

        # Create ReportFile instance
        report_file = ReportFile.objects.create(
            entry=entry,
            file=file_content,
            file_type=file_type,
            uploaded_by=report.created_by
        )

        return report_file
    except Exception as e:
        logger.error(f"File saving failed: {str(e)}")
        raise

def calculate_custom_field(calculated_field):
    # This is a simplified example. You might need to implement more complex logic
    # based on your specific requirements.
    report = calculated_field.report
    formula = calculated_field.formula
    
    # Example: Count of entries
    if formula == 'COUNT_ENTRIES':
        return report.entries.count()
    
    # Example: Average word count of entries
    elif formula == 'AVG_WORD_COUNT':
        word_counts = [len(entry.content.split()) for entry in report.entries.all()]
        return sum(word_counts) / len(word_counts) if word_counts else 0
    
    # Add more custom calculations as needed
    
    return None  # Return None if the formula is not recognized
